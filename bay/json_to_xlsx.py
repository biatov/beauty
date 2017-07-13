import json

from openpyxl import Workbook
from openpyxl.styles import Font
import re
from itertools import groupby
from nltk.corpus import stopwords

stop_words = stopwords.words('english')


def keywords(url, name):
    category = url.split('/')[3]
    if 'and' in category:
        key_url = category.split('and')
    elif 'makeup' in category:
        key_url = [category[:6], category[6:]]
    else:
        key_url = category
    key_name = list(filter(None, map(
        lambda i: i if re.search(r'\w', i) and not re.search(r'[\d\.,!\?;:…]', i) and i not in stop_words else '',
        name.lower().split())))
    key_words = list()
    if isinstance(key_url, str):
        key_words.append(key_url)
    else:
        key_words.extend(key_url)
    key_words.extend(key_name)
    key_words = [el for el, _ in groupby(key_words)]
    return ', '.join(key_words)


def read_xlsx(name, title):
    try:
        with open('bay/item/%s_items.json' % name) as f:
            urls = json.load(f)
    except FileNotFoundError:
        urls = list()
    image = list(map(lambda el: el, urls))

    wb = Workbook()
    wb.create_sheet(title)
    ws = wb[title]
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=2).value = 'Brand'
    ws.cell(row=1, column=3).font = Font(bold=True)
    ws.cell(row=1, column=3).value = 'Product Name'
    ws.cell(row=1, column=4).font = Font(bold=True)
    ws.cell(row=1, column=4).value = 'Product Description'
    ws.cell(row=1, column=5).font = Font(bold=True)
    ws.cell(row=1, column=5).value = 'Colour/Size'
    ws.cell(row=1, column=6).font = Font(bold=True)
    ws.cell(row=1, column=6).value = 'Meta tag'
    ws.cell(row=1, column=7).font = Font(bold=True)
    ws.cell(row=1, column=7).value = 'Price'
    ws.cell(row=1, column=8).font = Font(bold=True)
    ws.cell(row=1, column=8).value = 'Product URL'
    ws.cell(row=1, column=9).font = Font(bold=True)
    ws.cell(row=1, column=9).value = 'Image URL'

    for col, val in enumerate(image, start=2):
        ws.cell(row=col, column=1).value = val['id']
        ws.cell(row=col, column=2).value = val['brand']
        ws.cell(row=col, column=3).value = val['product_name']
        ws.cell(row=col, column=4).value = val['product_description']
        ws.cell(row=col, column=5).value = val['colour']
        ws.cell(row=col, column=6).value = keywords(val['product_url'], val['product_name'])
        ws.cell(row=col, column=7).value = val['price']
        ws.cell(row=col, column=8).value = val['product_url']
        ws.cell(row=col, column=9).value = val['image_url']
    try:
        wb.save("bay/excel/%s.xlsx" % name)
    except FileNotFoundError:
        pass


read_xlsx('nailcare', 'Nail')
# read_xlsx('accessories', 'Accessories')
