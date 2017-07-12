from openpyxl import Workbook
from openpyxl.styles import Font
import json



def read_xlsx():
    try:
        with open('nailcare_items.json') as f:
            urls = json.load(f)
    except FileNotFoundError:
        urls = list()
    image = list(map(lambda el: el, urls))

    wb = Workbook()
    wb.create_sheet('Nail')
    ws = wb['Nail']
    # wb.create_sheet('Electrical')
    # ws = wb['Electrical']
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).value = 'Brand'
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=2).value = 'Product Name'
    ws.cell(row=1, column=3).font = Font(bold=True)
    ws.cell(row=1, column=3).value = 'Product Description'
    ws.cell(row=1, column=4).font = Font(bold=True)
    ws.cell(row=1, column=4).value = 'Colour/Size'
    ws.cell(row=1, column=5).font = Font(bold=True)
    ws.cell(row=1, column=5).value = 'Meta tag'
    ws.cell(row=1, column=6).font = Font(bold=True)
    ws.cell(row=1, column=6).value = 'Price'

    for col, val in enumerate(image, start=2):
        ws.cell(row=col, column=1).value = val['brand']
        ws.cell(row=col, column=2).value = val['product_name']
        ws.cell(row=col, column=3).value = val['product_description']
        ws.cell(row=col, column=4).value = val['colour']
        ws.cell(row=col, column=5).value = val['meta_tag']
        ws.cell(row=col, column=6).value = val['price']
    try:
        wb.save("beautybay.xlsx")
    except FileNotFoundError:
        pass

read_xlsx()