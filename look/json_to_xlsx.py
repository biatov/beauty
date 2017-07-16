import json

from openpyxl import Workbook
from openpyxl.styles import Font
import re
from itertools import groupby
from nltk.corpus import stopwords

stop_words = stopwords.words('english')
stop_words_ch = ["、","。","〈","〉","《","》","一","一切","一则","一方面","一旦","一来","一样","一般","七","万一","三","上下","不仅","不但","不光","不单","不只","不如","不怕","不惟","不成","不拘","不比","不然","不特","不独","不管","不论","不过","不问","与","与其","与否","与此同时","且","两者","个","临","为","为了","为什么","为何","为着","乃","乃至","么","之","之一","之所以","之类","乌乎","乎","乘","九","也","也好","也罢","了","二","于","于是","于是乎","云云","五","人家","什么","什么样","从","从而","他","他人","他们","以","以便","以免","以及","以至","以至于","以致","们","任","任何","任凭","似的","但","但是","何","何况","何处","何时","作为","你","你们","使得","例如","依","依照","俺","俺们","倘","倘使","倘或","倘然","倘若","借","假使","假如","假若","像","八","六","兮","关于","其","其一","其中","其二","其他","其余","其它","其次","具体地说","具体说来","再者","再说","冒","冲","况且","几","几时","凭","凭借","则","别","别的","别说","到","前后","前者","加之","即","即令","即使","即便","即或","即若","又","及","及其","及至","反之","反过来","反过来说","另","另一方面","另外","只是","只有","只要","只限","叫","叮咚","可","可以","可是","可见","各","各个","各位","各种","各自","同","同时","向","向着","吓","吗","否则","吧","吧哒","吱","呀","呃","呕","呗","呜","呜呼","呢","呵","呸","呼哧","咋","和","咚","咦","咱","咱们","咳","哇","哈","哈哈","哉","哎","哎呀","哎哟","哗","哟","哦","哩","哪","哪个","哪些","哪儿","哪天","哪年","哪怕","哪样","哪边","哪里","哼","哼唷","唉","啊","啐","啥","啦","啪达","喂","喏","喔唷","嗡嗡","嗬","嗯","嗳","嘎","嘎登","嘘","嘛","嘻","嘿","四","因","因为","因此","因而","固然","在","在下","地","多","多少","她","她们","如","如上所述","如何","如其","如果","如此","如若","宁","宁可","宁愿","宁肯","它","它们","对","对于","将","尔后","尚且","就","就是","就是说","尽","尽管","岂但","己","并","并且","开外","开始","归","当","当着","彼","彼此","往","待","得","怎","怎么","怎么办","怎么样","怎样","总之","总的来看","总的来说","总的说来","总而言之","恰恰相反","您","慢说","我","我们","或","或是","或者","所","所以","打","把","抑或","拿","按","按照","换句话说","换言之","据","接着","故","故此","旁人","无宁","无论","既","既是","既然","时候","是","是的","替","有","有些","有关","有的","望","朝","朝着","本","本着","来","来着","极了","果然","果真","某","某个","某些","根据","正如","此","此外","此间","毋宁","每","每当","比","比如","比方","沿","沿着","漫说","焉","然则","然后","然而","照","照着","甚么","甚而","甚至","用","由","由于","由此可见","的","的话","相对而言","省得","着","着呢","矣","离","第","等","等等","管","紧接着","纵","纵令","纵使","纵然","经","经过","结果","给","继而","综上所述","罢了","者","而","而且","而况","而外","而已","而是","而言","能","腾","自","自个儿","自从","自各儿","自家","自己","自身","至","至于","若","若是","若非","莫若","虽","虽则","虽然","虽说","被","要","要不","要不是","要不然","要么","要是","让","论","设使","设若","该","诸位","谁","谁知","赶","起","起见","趁","趁着","越是","跟","较","较之","边","过","还是","还有","这","这个","这么","这么些","这么样","这么点儿","这些","这会儿","这儿","这就是说","这时","这样","这边","这里","进而","连","连同","通过","遵照","那","那个","那么","那么些","那么样","那些","那会儿","那儿","那时","那样","那边","那里","鄙人","鉴于","阿","除","除了","除此之外","除非","随","随着","零","非但","非徒","靠","顺","顺着","首先","︿","！","＃","＄","％","＆","（","）","＊","＋","，","０","１","２","３","４","５","６","７","８","９","：","；","＜","＞","？","＠","［","］","｛","｜","｝","～","￥"]
stop_words.extend(stop_words_ch)


def keywords(key_url, name):
    key_name = list(filter(None, map(
        lambda i: i.replace('+', '').replace('(', '').replace(')', '') if re.search(r'\w', i) and not re.search(r'[\d\.,!\?;:…]', i) and i not in stop_words else '',
        name.lower().split())))
    key_words = list()
    key_words.append(key_url)
    key_words.extend(key_name)
    key_words = [el for el, _ in groupby(key_words)]
    return ', '.join(key_words)


def read_xlsx(name, title, key):
    try:
        with open('look/remake_json/%s_ch_items_remake.json' % name) as f:
            urls = json.load(f)
    except FileNotFoundError:
        urls = list()
    image = list(map(lambda el: el, urls))

    wb = Workbook()
    wb.create_sheet(title)
    ws = wb[title]
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=2).value = '品牌'
    ws.cell(row=1, column=3).font = Font(bold=True)
    ws.cell(row=1, column=3).value = '产品名称'
    ws.cell(row=1, column=4).font = Font(bold=True)
    ws.cell(row=1, column=4).value = '描述'
    ws.cell(row=1, column=5).font = Font(bold=True)
    ws.cell(row=1, column=5).value = '颜色/尺寸'
    ws.cell(row=1, column=6).font = Font(bold=True)
    ws.cell(row=1, column=6).value = '元标签'
    ws.cell(row=1, column=7).font = Font(bold=True)
    ws.cell(row=1, column=7).value = '市场价'
    ws.cell(row=1, column=8).font = Font(bold=True)
    ws.cell(row=1, column=8).value = '产品网址'
    ws.cell(row=1, column=9).font = Font(bold=True)
    ws.cell(row=1, column=9).value = '图片网址'

    count = 1
    for col, val in enumerate(image, start=2):
        ws.cell(row=col, column=1).value = count
        ws.cell(row=col, column=2).value = val['brand']
        ws.cell(row=col, column=3).value = val['name']
        ws.cell(row=col, column=4).value = val['product_description']
        ws.cell(row=col, column=5).value = val['colour']
        ws.cell(row=col, column=6).value = keywords(key, val['name'])
        ws.cell(row=col, column=7).value = val['price']
        ws.cell(row=col, column=8).value = val['product_url']
        ws.cell(row=col, column=9).value = val['image_url']
        count += 1
    try:
        wb.save("look/excel_ch_remake/%s.xlsx" % name)
    except FileNotFoundError:
        pass

read_xlsx('hair', '护发', '护发')
read_xlsx('makeup', '彩妆', '彩妆')
read_xlsx('face', '护肤', '护肤')
read_xlsx('body', '护体', '护体')
read_xlsx('men', '男士', '男士')
read_xlsx('electrical', '美容电器', '美容电器')
