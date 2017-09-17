import json

from openpyxl import Workbook
import re
from itertools import groupby
from nltk.corpus import stopwords

stop_words = stopwords.words('english')


def keywords(catalogs, name):
    category = catalogs.replace('and', ',').replace('&', ',')
    key_cat = category.split(',')
    if isinstance(key_cat, list):
        key_cat = list(map(lambda j: j.lower().strip(), key_cat))
    else:
        key_cat = key_cat.lower()
    key_name = list(filter(None, map(
        lambda i: i if re.search(r'\w', i) and not re.search(r'[\d\.,!\?;:…]', i) and i not in stop_words else '',
        name.lower().split())))
    key_words = list()
    if isinstance(key_cat, str):
        key_words.append(key_cat)
    else:
        key_words.extend(key_cat)
    key_words.extend(key_name)
    new_words = list()
    for each in key_words:
        if each not in new_words:
            new_words.append(each)
    return ', '.join(new_words)


def read_xlsx(name, title):
    try:
        with open('ulta/full_data/%s/%s.json' % (name, name)) as f:
            urls = json.load(f)
    except FileNotFoundError:
        urls = list()
    image = list(map(lambda el: el, urls))

    wb = Workbook()
    wb.create_sheet(title)
    ws = wb[title]
    ws.cell(row=1, column=1).value = 'Product ID'
    ws.cell(row=1, column=2).value = 'Active (0/1)'
    ws.cell(row=1, column=3).value = 'Name *'
    ws.cell(row=1, column=4).value = 'Categories (x,y,z...)'
    ws.cell(row=1, column=5).value = 'Price tax included'
    ws.cell(row=1, column=6).value = 'Reference #'
    ws.cell(row=1, column=7).value = 'Quantity'
    ws.cell(row=1, column=8).value = 'Visibility'
    ws.cell(row=1, column=9).value = 'Short description'
    ws.cell(row=1, column=10).value = 'Description'
    ws.cell(row=1, column=11).value = 'Meta title'
    ws.cell(row=1, column=12).value = 'Meta keywords'
    ws.cell(row=1, column=13).value = 'Meta description'
    ws.cell(row=1, column=14).value = 'Image URLs (x,y,z...)'
    ws.cell(row=1, column=15).value = 'Feature(Name:Value:Position)'
    ws.cell(row=1, column=16).value = 'Brand'
    ws.cell(row=1, column=17).value = 'Discount percent '
    ws.cell(row=1, column=18).value = 'Discount from (yyyy-mm-dd)'
    ws.cell(row=1, column=19).value = 'Discount to (yyyy-mm-dd)'
    ws.cell(row=1, column=20).value = 'Tags (x,y,z…'
    ws.cell(row=1, column=21).value = 'On sale (0/1)'

    for col, val in enumerate(image, start=2):
        ws.cell(row=col, column=1).value = ''
        ws.cell(row=col, column=2).value = ''
        if val['sub_title']:
            ws.cell(row=col, column=3).value = '%s (%s)' % (val['title'], val['sub_title'])
        else:
            ws.cell(row=col, column=3).value = val['title']
        ws.cell(row=col, column=4).value = val['catalogs']
        ws.cell(row=col, column=5).value = val['price']
        ws.cell(row=col, column=6).value = ''
        ws.cell(row=col, column=7).value = ''
        ws.cell(row=col, column=8).value = ''
        ws.cell(row=col, column=9).value = ''
        ws.cell(row=col, column=10).value = val['description']
        ws.cell(row=col, column=11).value = ''
        ws.cell(row=col, column=12).value = keywords(val['catalogs'], val['title'])
        ws.cell(row=col, column=13).value = ''
        ws.cell(row=col, column=14).value = val['image']
        ws.cell(row=col, column=15).value = ''
        ws.cell(row=col, column=16).value = val['brand']
        ws.cell(row=col, column=17).value = ''
        ws.cell(row=col, column=18).value = ''
        ws.cell(row=col, column=19).value = ''
        ws.cell(row=col, column=20).value = ''
        ws.cell(row=col, column=21).value = ''

    try:
        wb.save("ulta/excel/%s.xlsx" % name)
    except FileNotFoundError:
        pass

# read_xlsx('hair', 'Hair')
# read_xlsx('makeup', 'Makeup')
# read_xlsx('nails', 'Nails')
# read_xlsx('skincare', 'Skincare')
# read_xlsx('toolsandbrushes', 'Tools and Brushes')
# read_xlsx('fragrance', 'Fragrance')
# read_xlsx('men', 'Men')
read_xlsx('bathandbody', 'Bath and Body')
